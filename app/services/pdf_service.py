"""
PDF/Excel出力サービス
帳簿・レポート・申告書類をPDF/Excel形式で出力
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io
import os
from typing import Dict, List
import logging
from datetime import datetime
import pandas as pd

logger = logging.getLogger(__name__)


class PDFService:
    """PDF生成サービス"""

    def __init__(self):
        self.font_registered = False
        self._register_fonts()

    def _register_fonts(self):
        """日本語フォント登録"""
        if self.font_registered:
            return

        # 日本語フォントのパスを探索
        font_paths = [
            "/usr/share/fonts/ipa-gothic/ipag.ttf",
            "/usr/share/fonts/truetype/ipa-gothic/ipag.ttf",
            "/usr/share/fonts/opentype/ipaexfont-gothic/ipaexg.ttf",
            "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
        ]

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont("IPAGothic", font_path))
                    self.font_registered = True
                    logger.info(f"Registered font: {font_path}")
                    return
                except Exception as e:
                    logger.warning(f"Failed to register font {font_path}: {e}")

        logger.warning("No Japanese font found, using default")

    def _get_font_name(self):
        """使用するフォント名を取得"""
        return "IPAGothic" if self.font_registered else "Helvetica"

    def generate_journal_pdf(
        self, journal_data: List[Dict], period: str, title: str = "仕訳帳"
    ) -> bytes:
        """仕訳帳PDF生成"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm
        )

        elements = []
        font_name = self._get_font_name()

        # タイトル
        title_style = ParagraphStyle(
            "Title",
            fontName=font_name,
            fontSize=16,
            alignment=1,
            spaceAfter=20,
        )
        elements.append(Paragraph(f"{title} ({period})", title_style))

        # 作成日時
        date_style = ParagraphStyle(
            "Date", fontName=font_name, fontSize=10, alignment=2
        )
        elements.append(
            Paragraph(
                f"作成日: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style
            )
        )
        elements.append(Spacer(1, 10 * mm))

        # テーブルデータ
        table_data = [["日付", "借方科目", "借方金額", "貸方科目", "貸方金額", "摘要"]]
        for row in journal_data:
            table_data.append(
                [
                    str(row.get("日付", "")),
                    str(row.get("借方科目", "")),
                    f"{row.get('借方金額', 0):,.0f}",
                    str(row.get("貸方科目", "")),
                    f"{row.get('貸方金額', 0):,.0f}",
                    str(row.get("摘要", ""))[:20],
                ]
            )

        table = Table(table_data, colWidths=[25 * mm, 30 * mm, 25 * mm, 30 * mm, 25 * mm, 35 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                    ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]
            )
        )

        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    def generate_profit_loss_pdf(self, pl_data: Dict) -> bytes:
        """損益計算書PDF生成"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)

        elements = []
        font_name = self._get_font_name()

        # タイトル
        title_style = ParagraphStyle(
            "Title", fontName=font_name, fontSize=16, alignment=1, spaceAfter=20
        )
        elements.append(
            Paragraph(f"損益計算書 ({pl_data.get('period', '')})", title_style)
        )

        # データテーブル
        table_data = [
            ["科目", "金額"],
            ["売上高", f"{pl_data.get('sales', 0):,.0f}円"],
            ["売上原価", f"{pl_data.get('cost_of_sales', 0):,.0f}円"],
            ["売上総利益", f"{pl_data.get('gross_profit', 0):,.0f}円"],
            ["", ""],
            ["【経費】", ""],
        ]

        for account, amount in pl_data.get("expenses_detail", {}).items():
            table_data.append([f"  {account}", f"{amount:,.0f}円"])

        table_data.extend(
            [
                ["経費合計", f"{pl_data.get('total_expenses', 0):,.0f}円"],
                ["", ""],
                ["営業利益", f"{pl_data.get('operating_profit', 0):,.0f}円"],
                ["利益率", f"{pl_data.get('profit_margin', 0):.1f}%"],
            ]
        )

        table = Table(table_data, colWidths=[100 * mm, 60 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("LINEBELOW", (0, 0), (-1, 0), 1, colors.black),
                    ("LINEBELOW", (0, 3), (-1, 3), 0.5, colors.grey),
                    ("LINEABOVE", (0, -2), (-1, -2), 1, colors.black),
                ]
            )
        )

        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()


class ExcelService:
    """Excel生成サービス"""

    def __init__(self):
        self.header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_journal_excel(
        self, journal_data: List[Dict], period: str
    ) -> bytes:
        """仕訳帳Excel生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"仕訳帳_{period}"

        # ヘッダー
        headers = ["日付", "借方科目", "借方金額", "貸方科目", "貸方金額", "摘要", "税区分"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # データ
        for row_num, row in enumerate(journal_data, 2):
            ws.cell(row=row_num, column=1, value=row.get("日付", ""))
            ws.cell(row=row_num, column=2, value=row.get("借方科目", ""))
            ws.cell(row=row_num, column=3, value=row.get("借方金額", 0))
            ws.cell(row=row_num, column=4, value=row.get("貸方科目", ""))
            ws.cell(row=row_num, column=5, value=row.get("貸方金額", 0))
            ws.cell(row=row_num, column=6, value=row.get("摘要", ""))
            ws.cell(row=row_num, column=7, value=row.get("税区分", ""))

            # 罫線
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = self.border

        # 金額列のフォーマット
        for row in range(2, len(journal_data) + 2):
            ws.cell(row=row, column=3).number_format = "#,##0"
            ws.cell(row=row, column=5).number_format = "#,##0"

        # 列幅調整
        column_widths = [12, 15, 12, 15, 12, 30, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_expense_summary_excel(
        self, summary_data: Dict, year_month: str
    ) -> bytes:
        """経費明細Excel生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"経費明細_{year_month}"

        # タイトル
        ws["A1"] = f"経費明細 {year_month}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        # ヘッダー
        ws["A3"] = "勘定科目"
        ws["B3"] = "金額"
        ws["A3"].font = self.header_font
        ws["A3"].fill = self.header_fill
        ws["B3"].font = self.header_font
        ws["B3"].fill = self.header_fill

        # データ
        row = 4
        total = 0
        for account, amount in summary_data.items():
            if account != "合計":
                ws.cell(row=row, column=1, value=account)
                ws.cell(row=row, column=2, value=amount)
                ws.cell(row=row, column=2).number_format = "#,##0"
                total += amount
                row += 1

        # 合計行
        ws.cell(row=row, column=1, value="合計")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = "#,##0"

        # 列幅
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_trial_balance_excel(
        self, trial_balance_df: pd.DataFrame, as_of_date: str
    ) -> bytes:
        """残高試算表Excel生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "残高試算表"

        # タイトル
        ws["A1"] = f"残高試算表 {as_of_date}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")

        # ヘッダー
        headers = ["勘定科目", "借方合計", "貸方合計", "借方残高", "貸方残高"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")

        # データ
        for row_num, (_, row) in enumerate(trial_balance_df.iterrows(), 4):
            ws.cell(row=row_num, column=1, value=row.get("勘定科目", ""))
            ws.cell(row=row_num, column=2, value=row.get("借方合計", 0))
            ws.cell(row=row_num, column=3, value=row.get("貸方合計", 0))
            ws.cell(row=row_num, column=4, value=row.get("借方残高", 0))
            ws.cell(row=row_num, column=5, value=row.get("貸方残高", 0))

            for col in range(2, 6):
                ws.cell(row=row_num, column=col).number_format = "#,##0"

        # 列幅
        ws.column_dimensions["A"].width = 20
        for col in ["B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()


pdf_service = PDFService()
excel_service = ExcelService()
